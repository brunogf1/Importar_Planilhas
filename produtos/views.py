from django.shortcuts import render, redirect, get_list_or_404
from openpyxl import load_workbook
from .models import Produto
from .forms import UploadExcelForm
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.template.loader import render_to_string

# Mapear os cabeçalhos do Excel para os campos do model Produto
CAMPOS_IMPORTAR = {
    'Loja': 'loja',
    'Endereço1': 'endereco1',
    'Cidade': 'cidade',
    'Uniformes': 'uniformes',
    'Vol. Unif.': 'vol_unif',
    'Cam./bone': 'cam_bone',
    'Vol. C / B': 'vol_cb',
    'Vol.': 'vol',
    'Peso': 'peso',
    'Nota': 'nota',
    'Emite NF': 'emite_nf',
}

def importar_produtos(request):
    form = UploadExcelForm()

    if request.method == 'POST':
        form = UploadExcelForm(request.POST, request.FILES)
        if form.is_valid():
            arquivo_excel = request.FILES['arquivo_excel']
            wb = load_workbook(arquivo_excel, data_only=True)
            sheet = wb.active

            Produto.objects.all().delete()

            cabecalho = [cell.value for cell in sheet[1]]

            # Mapeia campo_model --> índice na planilha
            campos_indices = {}
            for idx, col_name in enumerate(cabecalho):
                if col_name in CAMPOS_IMPORTAR:
                    campos_indices[CAMPOS_IMPORTAR[col_name]] = idx

            campos_model_necessarios = set(CAMPOS_IMPORTAR.values())
            campos_faltando = campos_model_necessarios - set(campos_indices.keys())
            if campos_faltando:
                return render(request, 'produtos/importar_produtos.html', {
                    'form': form,
                    'erro': f"Colunas obrigatórias não encontradas: {', '.join(campos_faltando)}"
                })

            for row in sheet.iter_rows(min_row=2, values_only=True):
                kwargs = {}
                for campo, idx in campos_indices.items():
                    kwargs[campo] = row[idx]
                Produto.objects.create(**kwargs)

            return redirect('lista_produtos') 
        
    return render(request, 'produtos/importar_produtos.html', {'form': form})

def lista_produtos(request):
    produtos_list = Produto.objects.all()

    search_query = request.GET.get('search', '')
    if search_query:
        produtos_list = produtos_list.filter(
            loja__icontains=search_query
        )

    sort_by = request.GET.get('sort', 'loja')
    order_direction = request.GET.get('direction', 'asc')

    if order_direction == 'desc':
        sort_by = f'-{sort_by}'

    produtos_list = produtos_list.order_by(sort_by)

    paginator = Paginator(produtos_list, 150)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'sort_by': sort_by.lstrip('-'),
        'order_direction': order_direction,
    }

    return render(request, 'produtos/lista_produtos.html', context)

def imprimir_produtos(request):
    if request.method == 'POST':
        produtos_ids = request.POST.getlist('produtos_selecionados')
        if produtos_ids:
            produtos = Produto.objects.filter(id__in=produtos_ids)
            html_content = render_to_string('produtos/imprimir_produtos.html', {'produtos': produtos})
            return render(request, 'produtos/imprimir_produtos.html', {'produtos': produtos})

    return redirect('lista_produtos')